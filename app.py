from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, render_template_string, make_response, session, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from datetime import datetime, date, timedelta
import os
from dotenv import load_dotenv
import io
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
import pandas as pd  # Import pandas
import openpyxl
import xlrd
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps, lru_cache
from pyotp import random_base32, TOTP
import traceback
import psycopg2
from psycopg2.extras import DictCursor
import time

# Load environment variables
load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', os.urandom(24))

# Database configuration - check environment
if os.getenv('FLASK_ENV') == 'production':
    # Use Neon PostgreSQL in production
    database_url = os.getenv('SQLALCHEMY_DATABASE_URI') or os.getenv('DATABASE_URL')
    if database_url and database_url.startswith('postgres://'):
        database_url = database_url.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
else:
    # Use SQLite in development
    basedir = os.path.abspath(os.path.dirname(__file__))
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'app.db')

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Cache configuration
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 31536000  # 1 year
app.config['STATIC_FOLDER'] = 'static'

# Cache decorators
def cache_for(seconds):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            cache_key = f.__name__ + str(args) + str(kwargs)
            cached_result = cache.get(cache_key)
            if cached_result is not None:
                return cached_result
            result = f(*args, **kwargs)
            cache.set(cache_key, result, timeout=seconds)
            return result
        return decorated_function
    return decorator

# Simple in-memory cache
class SimpleCache:
    def __init__(self):
        self._cache = {}
        
    def get(self, key):
        if key in self._cache:
            item = self._cache[key]
            if item['expires'] > datetime.utcnow():
                return item['value']
            else:
                del self._cache[key]
        return None
        
    def set(self, key, value, timeout=300):
        self._cache[key] = {
            'value': value,
            'expires': datetime.utcnow() + timedelta(seconds=timeout)
        }
        
    def delete(self, key):
        if key in self._cache:
            del self._cache[key]

cache = SimpleCache()

db = SQLAlchemy(app)
migrate = Migrate(app, db)

# Models
class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    item_code = db.Column(db.String(20), unique=True, nullable=False)
    description = db.Column(db.String(200), nullable=False)
    tamil_name = db.Column(db.String(200))  # Optional Tamil name
    uom = db.Column(db.String(10), nullable=False)
    price = db.Column(db.Float, nullable=False)
    stock = db.Column(db.Integer, default=0)
    restock_level = db.Column(db.Integer, default=0)  # Level at which to restock
    stock_locations = db.Column(db.String(500))  # Comma-separated location tags
    tags = db.Column(db.String(500))  # Comma-separated tags
    notes = db.Column(db.Text)  # Product notes

    @property
    def serialize(self):
        return {
            'id': self.id,
            'item_code': self.item_code,
            'description': self.description,
            'tamil_name': self.tamil_name,
            'uom': self.uom,
            'price': self.price,
            'stock': self.stock,
            'restock_level': self.restock_level,
            'stock_locations': self.stock_locations,
            'tags': self.tags,
            'notes': self.notes
        }

class Customer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    phone = db.Column(db.String(20))
    email = db.Column(db.String(120))
    address = db.Column(db.Text)
    balance = db.Column(db.Float, default=0.0)  # Positive means customer owes money, negative means surplus
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    invoices = db.relationship('Invoice', backref='customer', lazy=True)
    transactions = db.relationship('CustomerTransaction', backref='customer', lazy=True)
    receivables = db.relationship('CustomerReceivable', backref='customer', lazy=True)

    def update_balance(self):
        # Start with 0 balance
        total_balance = 0.0
        
        # Add receivables (amounts customer owes)
        for receivable in self.receivables:
            total_balance += receivable.amount + receivable.additional_amount
        
        # Subtract payments (reduce balance) or add refunds (increase balance)
        for transaction in self.transactions:
            if transaction.transaction_type == 'payment':
                total_balance -= transaction.amount
            elif transaction.transaction_type == 'refund':
                total_balance += transaction.amount
        
        # Update the balance
        self.balance = round(total_balance, 2)
        db.session.commit()

class CustomerTransaction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)
    date = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    amount = db.Column(db.Float, nullable=False)
    transaction_type = db.Column(db.String(20), nullable=False)  # 'payment' or 'refund'
    payment_method = db.Column(db.String(50))  # 'cash', 'card', 'upi', etc.
    notes = db.Column(db.Text)
    reference_number = db.Column(db.String(50))  # For tracking payment references

class CustomerReceivable(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    notes = db.Column(db.Text, nullable=False)
    date = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    invoice_id = db.Column(db.Integer, db.ForeignKey('invoice.id'))  # Optional, for linked invoices
    additional_amount = db.Column(db.Float, default=0.0)  # For additional amounts on linked invoices
    
    # Remove the duplicate backref and use foreign_keys for clarity
    invoice = db.relationship('Invoice', backref='receivable')

class Invoice(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    order_number = db.Column(db.String(3), nullable=False)  # Daily 3-digit number
    date = db.Column(db.Date, nullable=False, default=date.today)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'))
    customer_name = db.Column(db.String(200))
    total_amount = db.Column(db.Float, default=0.0)
    total_items = db.Column(db.Integer, default=0)
    items = db.relationship('InvoiceItem', backref='invoice', lazy=True, cascade="all, delete-orphan")
    payment_status = db.Column(db.String(20), default='pending')  # pending, partial, paid

    @classmethod
    def generate_order_number(cls):
        # Get the latest invoice for today
        today = date.today()
        latest_invoice = cls.query.filter(
            func.date(cls.date) == today
        ).order_by(cls.id.desc()).first()
        
        if latest_invoice:
            last_number = int(latest_invoice.order_number)
            new_number = str(last_number + 1).zfill(3)
        else:
            new_number = '001'
        
        return new_number

class InvoiceItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    invoice_id = db.Column(db.Integer, db.ForeignKey('invoice.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    price = db.Column(db.Float, nullable=False)
    amount = db.Column(db.Float, nullable=False)
    product = db.relationship('Product', backref='invoice_items')

class PrintTemplate(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    type = db.Column(db.String(20), nullable=False)  # 'invoice' or 'summary'
    content = db.Column(db.Text, nullable=False)
    is_default = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

# Models
class Permission(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)
    description = db.Column(db.String(200))

# User-Permission association table
user_permissions = db.Table('user_permissions',
    db.Column('user_id', db.Integer, db.ForeignKey('user.id'), primary_key=True),
    db.Column('permission_id', db.Integer, db.ForeignKey('permission.id'), primary_key=True)
)

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(255), nullable=False)  # Increased length for hashed password
    role = db.Column(db.String(20), nullable=False, default='user')
    totp_secret = db.Column(db.String(32))
    totp_enabled = db.Column(db.Boolean, default=False)
    permissions = db.relationship('Permission', secondary=user_permissions, lazy='subquery',
        backref=db.backref('users', lazy=True))

    def verify_totp(self, token):
        if not self.totp_enabled or not self.totp_secret:
            return True
        totp = TOTP(self.totp_secret)
        return totp.verify(token)

    def has_permission(self, permission_name):
        if self.role == 'admin':  # Admin has all permissions
            return True
        return any(p.name == permission_name for p in self.permissions)

    def has_any_permission(self, permission_names):
        if self.role == 'admin':  # Admin has all permissions
            return True
        return any(p.name in permission_names for p in self.permissions)

# Function to initialize default permissions
def init_permissions():
    default_permissions = [
        ('view_customers', 'Can view customers list and details'),
        ('edit_customers', 'Can create, edit and delete customers'),
        ('view_invoices', 'Can view invoices'),
        ('create_invoices', 'Can create new invoices'),
        ('edit_invoices', 'Can edit existing invoices'),
        ('delete_invoices', 'Can delete invoices'),
        ('view_products', 'Can view products list'),
        ('edit_products', 'Can create, edit and delete products'),
        ('import_products', 'Can import products from Excel/CSV'),
        ('export_products', 'Can export products to Excel/CSV'),
        ('manage_stock', 'Can update product stock levels'),
        ('delete_products', 'Can delete products'),
        # Product field-specific permissions
        ('view_product_code', 'Can view product item codes'),
        ('edit_product_code', 'Can edit product item codes'),
        ('view_product_description', 'Can view product descriptions'),
        ('edit_product_description', 'Can edit product descriptions'),
        ('view_product_tamil', 'Can view product Tamil names'),
        ('edit_product_tamil', 'Can edit product Tamil names'),
        ('view_product_uom', 'Can view product UOM'),
        ('edit_product_uom', 'Can edit product UOM'),
        ('view_product_price', 'Can view product prices'),
        ('edit_product_price', 'Can edit product prices'),
        ('view_product_stock', 'Can view product stock levels'),
        ('view_invoice_stock', 'Can view stock in invoice search dropdown'),
        ('edit_product_stock', 'Can edit product stock levels'),
        ('view_product_restock', 'Can view product restock levels'),
        ('edit_product_restock', 'Can edit product restock levels'),
        ('view_product_locations', 'Can view product locations'),
        ('edit_product_locations', 'Can edit product locations'),
        ('view_product_tags', 'Can view product tags'),
        ('edit_product_tags', 'Can edit product tags'),
        ('view_product_notes', 'Can view product notes'),
        ('edit_product_notes', 'Can edit product notes'),
        ('view_product_suppliers', 'Can view product suppliers'),
        ('manage_settings', 'Can manage system settings'),
        ('manage_users', 'Can manage users'),
        ('view_suppliers', 'Can view suppliers list'),
        ('edit_suppliers', 'Can create, edit and delete suppliers')
    ]
    
    for name, description in default_permissions:
        if not Permission.query.filter_by(name=name).first():
            permission = Permission(name=name, description=description)
            db.session.add(permission)
    
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error initializing permissions: {str(e)}")

class Settings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    calculator_code = db.Column(db.String(20), default='9999')
    wallpaper_path = db.Column(db.String(200))  # Path to the wallpaper file

# Login required decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Permission required decorator
def permission_required(permission_name):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            
            user = User.query.get(session['user_id'])
            if not user:
                return redirect(url_for('login'))
            
            if not user.has_permission(permission_name):
                flash('You do not have permission to access this page')
                return redirect(url_for('index'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Admin required decorator
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        user = User.query.get(session['user_id'])
        if not user or user.role != 'admin':
            flash('Admin access required')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

# Add helper function for getting user info
def get_user(user_id):
    return User.query.get(user_id)

# Add to template context
app.jinja_env.globals.update(get_user=get_user)

@app.context_processor
def utility_processor():
    def check_permission(permission_name):
        if 'user_id' not in session:
            return False
        if session.get('is_admin', False):
            return True
        return permission_name in session.get('permissions', [])
    
    def get_user(user_id):
        return User.query.get(user_id)
    
    return {
        'check_permission': check_permission,
        'get_user': get_user
    }

# Create tables and add sample data
def init_db():
    with app.app_context():
        try:
            # Create all tables
            db.create_all()
            
            # Initialize permissions first
            init_permissions()
            
            # Create admin user if none exists
            admin_user = User.query.filter_by(role='admin').first()
            if not admin_user:
                admin_user = User(
                    username='admin',
                    password=generate_password_hash('admin'),
                    role='admin'
                )
                db.session.add(admin_user)
                db.session.commit()
                print("Admin user created successfully")

            # Initialize settings if not exists
            settings = Settings.query.first()
            if not settings:
                settings = Settings(calculator_code='9999')
                db.session.add(settings)
                db.session.commit()
                
        except Exception as e:
            db.session.rollback()
            print(f"Error in init_db: {str(e)}")
            raise

# Helper function to delete invoices and invoice items
def _delete_all_invoices(restore_stock=False):
    # Get all invoices and their items before deletion
    invoices = Invoice.query.all()
    
    if restore_stock:
        # Restore stock for all items
        for invoice in invoices:
            for item in invoice.items:
                item.product.stock += item.quantity
    
    InvoiceItem.query.delete()
    Invoice.query.delete()

# Helper function to delete products, invoices and invoice items
def _delete_all_products():
    InvoiceItem.query.delete()
    Invoice.query.delete()
    Product.query.delete()

# Routes
@app.route('/')
@login_required
def index():
    try:
        # Get current user
        user = User.query.get(session['user_id'])
        if not user:
            session.clear()
            return redirect(url_for('login'))

        # Get settings
        settings = Settings.query.first()
        wallpaper_url = url_for('static', filename=f'wallpapers/{settings.wallpaper_path}') if settings and settings.wallpaper_path else None

        # Initialize stats dictionary
        stats = {
            'total_products': 0,
            'total_invoices': 0,
            'total_sales': 0,
            'recent_invoices': [],
            'low_stock_products': [],
            'total_inventory_cost': 0,
            'stockout_count': 0
        }

        # Only show stats if user has appropriate permissions
        if user.role == 'admin' or user.has_any_permission([
            'view_products', 'view_invoices', 'view_product_stock'
        ]):
            try:
                # Get basic stats that most users should see
                if user.has_permission('view_products'):
                    stats['total_products'] = db.session.query(func.count(Product.id)).scalar() or 0

                if user.has_permission('view_invoices'):
                    stats['total_invoices'] = db.session.query(func.count(Invoice.id)).scalar() or 0
                    stats['total_sales'] = db.session.query(func.sum(Invoice.total_amount)).scalar() or 0
                    stats['recent_invoices'] = Invoice.query.order_by(Invoice.date.desc()).limit(5).all()

                # Stock-related stats for users with stock permissions
                if user.has_permission('view_product_stock'):
                    stats['low_stock_products'] = Product.query.filter(
                        Product.stock <= Product.restock_level
                    ).all()
                    stats['stockout_count'] = db.session.query(
                        func.count(Product.id)
                    ).filter(Product.stock == 0).scalar() or 0
                    
                    # Calculate total inventory cost
                    total_inventory = db.session.query(
                        func.sum(Product.stock * Product.price)
                    ).scalar()
                    stats['total_inventory_cost'] = total_inventory or 0

            except Exception as e:
                print(f"Error calculating stats: {str(e)}")
                # Continue with empty stats rather than failing completely

        # Render template
        return render_template('index.html', 
                            stats=stats, 
                            wallpaper_url=wallpaper_url,
                            user=user)
        
    except Exception as e:
        print(f"Error in index route: {str(e)}")
        db.session.rollback()
        return "An error occurred. Please try again.", 500

@app.route('/products/search')
@login_required
def search_products():
    query = request.args.get('q', '').lower()
    search_terms = [term.strip() for term in query.split() if term.strip()]
    
    if not search_terms:
        return jsonify([])
    
    # Build the search query
    search_query = Product.query
    
    # Search across multiple fields with OR conditions
    search_conditions = []
    for term in search_terms:
        term_conditions = []
        # Search in item_code
        term_conditions.append(Product.item_code.ilike(f'%{term}%'))
        # Search in description
        term_conditions.append(Product.description.ilike(f'%{term}%'))
        # Search in tamil_name
        term_conditions.append(Product.tamil_name.ilike(f'%{term}%'))
        # Search in stock_locations
        term_conditions.append(Product.stock_locations.ilike(f'%{term}%'))
        # Search in tags
        term_conditions.append(Product.tags.ilike(f'%{term}%'))
        # Search in notes
        term_conditions.append(Product.notes.ilike(f'%{term}%'))
        # Search in UOM
        term_conditions.append(Product.uom.ilike(f'%{term}%'))
        
        # Combine conditions for this term with OR
        search_conditions.append(db.or_(*term_conditions))
    
    # Apply all term conditions with AND
    search_query = search_query.filter(db.and_(*search_conditions))
    
    # Execute query and get results
    products = search_query.order_by(Product.item_code).all()
    
    # Return serialized results with additional fields
    return jsonify([{
        **product.serialize,
        'stock_locations_display': product.stock_locations or '',
        'tags_display': product.tags or '',
        'notes_display': product.notes or ''
    } for product in products])

@app.route('/products', methods=['GET', 'POST'])
@login_required
@permission_required('view_products')
def products():
    if request.method == 'POST':
        # Check if user has edit permission
        if not session.get('is_admin', False) and 'edit_products' not in session.get('permissions', []):
            return jsonify({'success': False, 'error': 'Permission denied'})
            
        data = request.json
        
        # Server-side validation
        if not all(key in data for key in ('item_code', 'description', 'uom', 'price')):
            return jsonify({'success': False, 'error': 'Missing required data'})
        
        try:
            # Convert and validate numeric fields
            try:
                price = float(data['price'])
                stock = int(data.get('stock', 0))
                restock_level = int(data.get('restock_level', 0))
            except ValueError as e:
                return jsonify({'success': False, 'error': f'Invalid numeric value: {str(e)}'})
            
            # Create new product
            product = Product(
                item_code=data['item_code'],
                description=data['description'],
                tamil_name=data.get('tamil_name', ''),
                uom=data['uom'],
                price=price,
                stock=stock,
                restock_level=restock_level,
                stock_locations=data.get('stock_locations', ''),
                tags=data.get('tags', ''),
                notes=data.get('notes', '')
            )
            
            db.session.add(product)
            db.session.commit()
            return jsonify({'success': True})
            
        except IntegrityError:
            db.session.rollback()
            return jsonify({'success': False, 'error': 'Item code already exists'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)})
    
    # GET request handling
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    
    # Generate cache key based on page and per_page
    cache_key = f'products_page_{page}_{per_page}'
    cached_data = cache.get(cache_key)
    
    if cached_data:
        return cached_data
    
    # Query with pagination
    pagination = Product.query.order_by(Product.item_code).paginate(
        page=page, 
        per_page=per_page,
        error_out=False
    )
    
    products = pagination.items
    total_pages = pagination.pages
    total_items = pagination.total
    
    response = render_template(
        'products.html',
        products=products,
        pagination=pagination,
        total_pages=total_pages,
        total_items=total_items,
        current_page=page
    )
    
    # Cache the response for 5 minutes
    cache.set(cache_key, response, timeout=300)
    
    return response

@app.route('/products/<int:id>', methods=['PUT'])
@login_required
def update_product(id):
    product = Product.query.get_or_404(id)
    data = request.json
    current_user = User.query.get(session['user_id'])
    
    try:
        # Check permissions for each field that is being updated
        if 'item_code' in data and not (current_user.role == 'admin' or current_user.has_permission('edit_product_code')):
            return jsonify({'success': False, 'error': 'Permission denied: Cannot edit item code'})
            
        if 'description' in data and not (current_user.role == 'admin' or current_user.has_permission('edit_product_description')):
            return jsonify({'success': False, 'error': 'Permission denied: Cannot edit description'})
            
        if 'tamil_name' in data and not (current_user.role == 'admin' or current_user.has_permission('edit_product_tamil')):
            return jsonify({'success': False, 'error': 'Permission denied: Cannot edit Tamil name'})
            
        if 'uom' in data and not (current_user.role == 'admin' or current_user.has_permission('edit_product_uom')):
            return jsonify({'success': False, 'error': 'Permission denied: Cannot edit UOM'})
            
        if 'price' in data and not (current_user.role == 'admin' or current_user.has_permission('edit_product_price')):
            return jsonify({'success': False, 'error': 'Permission denied: Cannot edit price'})
            
        if 'stock' in data and not (current_user.role == 'admin' or current_user.has_permission('edit_product_stock')):
            return jsonify({'success': False, 'error': 'Permission denied: Cannot edit stock'})
            
        if 'restock_level' in data and not (current_user.role == 'admin' or current_user.has_permission('edit_product_restock')):
            return jsonify({'success': False, 'error': 'Permission denied: Cannot edit restock level'})
            
        if 'stock_locations' in data and not (current_user.role == 'admin' or current_user.has_permission('edit_product_locations')):
            return jsonify({'success': False, 'error': 'Permission denied: Cannot edit locations'})
            
        if 'tags' in data and not (current_user.role == 'admin' or current_user.has_permission('edit_product_tags')):
            return jsonify({'success': False, 'error': 'Permission denied: Cannot edit tags'})
            
        if 'notes' in data and not (current_user.role == 'admin' or current_user.has_permission('edit_product_notes')):
            return jsonify({'success': False, 'error': 'Permission denied: Cannot edit notes'})
        
        # Update only the fields that are present in the request data and user has permission for
        if 'item_code' in data and (current_user.role == 'admin' or current_user.has_permission('edit_product_code')):
            product.item_code = data['item_code']
            
        if 'description' in data and (current_user.role == 'admin' or current_user.has_permission('edit_product_description')):
            product.description = data['description']
            
        if 'tamil_name' in data and (current_user.role == 'admin' or current_user.has_permission('edit_product_tamil')):
            product.tamil_name = data['tamil_name']
            
        if 'uom' in data and (current_user.role == 'admin' or current_user.has_permission('edit_product_uom')):
            product.uom = data['uom']
            
        if 'price' in data and (current_user.role == 'admin' or current_user.has_permission('edit_product_price')):
            product.price = float(data['price'])
            
        if 'stock' in data and (current_user.role == 'admin' or current_user.has_permission('edit_product_stock')):
            product.stock = int(data['stock'])
            
        if 'restock_level' in data and (current_user.role == 'admin' or current_user.has_permission('edit_product_restock')):
            product.restock_level = int(data['restock_level'])
            
        if 'stock_locations' in data and (current_user.role == 'admin' or current_user.has_permission('edit_product_locations')):
            product.stock_locations = data['stock_locations']
            
        if 'tags' in data and (current_user.role == 'admin' or current_user.has_permission('edit_product_tags')):
            product.tags = data['tags']
            
        if 'notes' in data and (current_user.role == 'admin' or current_user.has_permission('edit_product_notes')):
            product.notes = data['notes']
        
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/products/<int:id>', methods=['DELETE'])
@login_required
@permission_required('delete_products')
def delete_product(id):
    product = Product.query.get_or_404(id)
    
    try:
        # Check if product has any invoice items
        if product.invoice_items:
            return jsonify({
                'success': False, 
                'error': 'Cannot delete product with existing invoices. Please delete related invoices first.'
            }), 400
            
        db.session.delete(product)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/invoices')
@login_required
@permission_required('view_invoices')
def invoices():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = Invoice.query
    if start_date and end_date:
        query = query.filter(
            Invoice.date >= datetime.strptime(start_date, '%Y-%m-%d').date(),
            Invoice.date <= datetime.strptime(end_date, '%Y-%m-%d').date()
        )
    
    invoices = query.order_by(Invoice.date.desc()).all()
    return render_template('invoices.html', invoices=invoices)

@app.route('/invoices/<int:id>', methods=['GET', 'PUT', 'DELETE'])
def invoice(id):
    invoice = Invoice.query.get_or_404(id)
    
    if request.method == 'GET':
        return jsonify({
            'id': invoice.id,
            'order_number': invoice.order_number,
            'date': invoice.date.isoformat(),
            'customer_name': invoice.customer_name,
            'total_amount': invoice.total_amount,
            'total_items': invoice.total_items,
            'items': [{
                'id': item.id,
                'product_id': item.product_id,
                'product_code': item.product.item_code,
                'description': item.product.description,
                'uom': item.product.uom,
                'quantity': item.quantity,
                'price': item.price,
                'amount': item.amount
            } for item in invoice.items]
        })
    
    elif request.method == 'PUT':
        data = request.json
        invoice.date = datetime.strptime(data['date'], '%Y-%m-%d').date()
        invoice.customer_name = data['customer_name']
        invoice.total_amount = float(data['total_amount'])
        invoice.total_items = int(data['total_items'])
        
        # Restore stock for existing items
        for item in invoice.items:
            item.product.stock += item.quantity
        
        # Remove existing items
        for item in invoice.items:
            db.session.delete(item)
        
        # Add new items
        for item_data in data['items']:
            product = Product.query.get(item_data['product_id'])
            # Remove stock validation check
            item = InvoiceItem(
                invoice=invoice,
                product_id=item_data['product_id'],
                quantity=item_data['quantity'],
                price=item_data['price'],
                amount=item_data['amount']
            )
            product.stock -= item_data['quantity']
            db.session.add(item)
        
        try:
            db.session.commit()
            # Return the invoice ID so the frontend can handle printing
            return jsonify({
                'success': True,
                'id': invoice.id,
                'redirect': url_for('new_invoice')
            })
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)})
    
    elif request.method == 'DELETE':
        data = request.json
        restore_stock = data.get('restore_stock', False)
        
        try:
            if restore_stock:
                # Restore stock for all items
                for item in invoice.items:
                    item.product.stock += item.quantity
            
            db.session.delete(invoice)
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)})

@app.route('/invoices/<int:id>/print')
def print_invoice(id):
    invoice = Invoice.query.get_or_404(id)
    return render_template('print_invoice.html', invoice=invoice)

@app.route('/invoices/<int:id>/print-tamil')
def print_invoice_tamil(id):
    invoice = Invoice.query.get_or_404(id)
    return render_template('print_invoice_tamil.html', invoice=invoice)

@app.route('/new_invoice', methods=['GET', 'POST'])
def new_invoice():
    if request.method == 'POST':
        data = request.json
        invoice = Invoice(
            order_number=Invoice.generate_order_number(),
            date=datetime.strptime(data['date'], '%Y-%m-%d').date(),
            customer_id=data.get('customer_id'),
            customer_name=data['customer_name'],
            total_amount=float(data['total_amount']),
            total_items=int(data['total_items'])
        )
        
        try:
            db.session.add(invoice)
            
            for item_data in data['items']:
                product = Product.query.get(item_data['product_id'])
                invoice_item = InvoiceItem(
                    invoice=invoice,
                    product_id=item_data['product_id'],
                    quantity=item_data['quantity'],
                    price=item_data['price'],
                    amount=item_data['amount']
                )
                product.stock -= item_data['quantity']
                db.session.add(invoice_item)
            
            if invoice.customer_id:
                customer = Customer.query.get(invoice.customer_id)
                customer.update_balance()
            
            db.session.commit()
            # Change the response to include redirect URL
            return jsonify({
                'success': True, 
                'id': invoice.id,
                'redirect': url_for('new_invoice')
            })
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)})
    
    customer_id = request.args.get('customer_id')
    customer = Customer.query.get(customer_id) if customer_id else None
    products = Product.query.all()
    return render_template('new_invoice.html', products=products, customer=customer)

@app.route('/invoices/print_summary')
@login_required
def print_summary():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = Invoice.query
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        query = query.filter(Invoice.date >= start_date)
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        query = query.filter(Invoice.date <= end_date)
    
    invoices = query.order_by(Invoice.date).all()
    total_amount = sum(invoice.total_amount for invoice in invoices)
    
    return render_template('print_summary.html',
                         invoices=invoices,
                         total_amount=total_amount)

@app.route('/products/import', methods=['POST'])
@login_required
@permission_required('import_products')
def import_products():
    try:
        if 'file' not in request.files:
            print("No file in request.files")
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400

        file = request.files['file']
        if file.filename == '':
            print("Empty filename")
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            print(f"Invalid file type: {file.filename}")
            return jsonify({'success': False, 'error': 'Invalid file format. Please upload an Excel file.'}), 400
        
        print(f"Processing file: {file.filename}")
        
        # Initialize counters and error list
        total_count = 0
        imported_count = 0
        error_details = []
        products = []
        batch_size = 20
        current_batch = []
        batch_number = 0
        
        try:
            # Read the Excel file
            wb = openpyxl.load_workbook(file, data_only=True)
            sheet = wb.active
            
            # Get total number of rows for progress calculation
            total_rows = sum(1 for row in sheet.iter_rows(min_row=2) if any(cell.value for cell in row))
            
            # Validate header row
            header_row = next(sheet.iter_rows(min_row=1, max_row=1))
            if len(header_row) < 6:
                print("Invalid header row")
                return jsonify({'success': False, 'error': 'Invalid file format. Missing required columns.'}), 400
            
            # Process data rows
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                total_count += 1
                try:
                    # Skip empty rows
                    if not any(cell.value for cell in row):
                        continue
                        
                    # Basic validation
                    if not row[0].value or not row[1].value or not row[3].value:
                        error_details.append(f"Row {row_idx}: Missing required fields (Item Code, Description, or UOM)")
                        continue
                    
                    # Handle price and stock
                    try:
                        price = float(str(row[4].value or '0').replace(',', ''))
                        stock = int(float(str(row[5].value or '0').replace(',', '')))
                    except (ValueError, TypeError) as e:
                        error_details.append(f"Row {row_idx}: Invalid price or stock value")
                        print(f"Error parsing price/stock in row {row_idx}: {str(e)}")
                        continue
                    
                    product = {
                        'item_code': str(row[0].value).strip(),
                        'description': str(row[1].value).strip(),
                        'tamil_name': str(row[2].value).strip() if row[2].value else None,
                        'uom': str(row[3].value).strip(),
                        'price': price,
                        'stock': stock,
                        'restock_level': int(float(str(row[6].value or '0').replace(',', ''))),
                        'stock_locations': str(row[7].value).strip() if len(row) > 7 and row[7].value else None,
                        'tags': str(row[8].value).strip() if len(row) > 8 and row[8].value else None,
                        'notes': str(row[9].value).strip() if len(row) > 9 and row[9].value else None
                    }
                    current_batch.append(product)
                    
                    # Process batch when it reaches batch_size
                    if len(current_batch) >= batch_size:
                        batch_number += 1
                        batch_result = process_product_batch(current_batch, error_details, batch_number)
                        imported_count += batch_result
                        current_batch = []
                        time.sleep(1)  # Add a 1-second delay between batches
                        
                except Exception as e:
                    error_details.append(f"Row {row_idx}: {str(e)}")
                    print(f"Error processing row {row_idx}: {str(e)}")
                    continue
            
            # Process remaining products in the last batch
            if current_batch:
                batch_number += 1
                imported_count += process_product_batch(current_batch, error_details, batch_number)
            
            response_data = {
                'success': imported_count > 0,
                'message': f"Successfully imported {imported_count} products.",
                'total_count': total_count,
                'imported_count': imported_count,
                'error_count': len(error_details),
                'error_details': error_details,
                'total_batches': batch_number
            }
            print(f"Import completed: {response_data}")
            return jsonify(response_data), 200 if imported_count > 0 else 400
            
        except Exception as e:
            print(f"Error reading Excel file: {str(e)}")
            return jsonify({
                'success': False,
                'error': f"Error reading Excel file: {str(e)}",
                'total_count': total_count,
                'imported_count': imported_count,
                'error_count': len(error_details),
                'error_details': error_details
            }), 400
            
    except Exception as e:
        print(f"Unexpected error: {str(e)}")
        return jsonify({
            'success': False,
            'error': f"Import failed: {str(e)}",
            'details': traceback.format_exc()
        }), 500

def process_product_batch(batch, error_details, batch_number):
    """Process a batch of products and return the number of successfully imported products"""
    imported_count = 0
    
    print(f"Processing batch {batch_number} with {len(batch)} products")
    
    for product_data in batch:
        try:
            # Validate item_code format
            if not product_data['item_code'] or len(product_data['item_code']) > 20:
                error_details.append(f"Invalid item code format: {product_data['item_code']}")
                continue
            
            product = Product.query.filter_by(item_code=product_data['item_code']).first()
            if product:
                # Update existing product
                for key, value in product_data.items():
                    setattr(product, key, value)
            else:
                # Create new product
                product = Product(**product_data)
                db.session.add(product)
            
            db.session.commit()
            imported_count += 1
            print(f"Imported product: {product_data['item_code']}")
        except Exception as e:
            db.session.rollback()
            error_details.append(f"Error with item code {product_data['item_code']}: {str(e)}")
            print(f"Error importing product {product_data['item_code']}: {str(e)}")
            continue
    
    print(f"Completed batch {batch_number}: {imported_count} products imported successfully")
    return imported_count

@app.route('/products/export')
@login_required
@permission_required('export_products')
def export_products():
    products = Product.query.all()
    
    # Create DataFrame with all fields
    df = pd.DataFrame([{
        'Item Code': p.item_code,
        'Description': p.description,
        'Tamil Name': p.tamil_name,
        'UOM': p.uom,
        'Price': p.price,
        'Stock': p.stock,
        'Restock Level': p.restock_level,
        'Stock Locations': p.stock_locations,
        'Tags': p.tags,
        'Notes': p.notes
    } for p in products])

    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    workbook = writer.book
    
    # Add formats
    header_format = workbook.add_format({
        'bold': True,
        'text_wrap': True,
        'valign': 'top',
        'bg_color': '#D9D9D9',
        'border': 1
    })
    
    # Write DataFrame to Excel
    df.to_excel(writer, sheet_name='Products', index=False)
    
    # Get the worksheet object
    worksheet = writer.sheets['Products']
    
    # Format the header row
    for col_num, value in enumerate(df.columns.values):
        worksheet.write(0, col_num, value, header_format)
        worksheet.set_column(col_num, col_num, 15)  # Set column width
    
    # Set specific column widths
    worksheet.set_column('B:B', 30)  # Description column
    worksheet.set_column('C:C', 20)  # Tamil Name column
    worksheet.set_column('H:H', 25)  # Stock Locations column
    worksheet.set_column('I:I', 20)  # Tags column
    worksheet.set_column('J:J', 30)  # Notes column
    
    writer.close()
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='products.xlsx'
    )


@app.route('/settings/print-templates')
def print_templates():
    invoice_templates = PrintTemplate.query.filter_by(type='invoice').all()
    summary_templates = PrintTemplate.query.filter_by(type='summary').all()
    return render_template('settings/print_templates.html',
                         invoice_templates=invoice_templates,
                         summary_templates=summary_templates)

@app.route('/settings/print-templates/new', methods=['GET', 'POST'])
@login_required
@admin_required
def new_print_template():
    if request.method == 'POST':
        data = request.json
        template = PrintTemplate(
            name=data['name'],
            type=data['type'],
            content=data['content'],
            is_default=data.get('is_default', False)
        )
        
        if template.is_default:
            # Remove default flag from other templates of same type
            PrintTemplate.query.filter_by(type=template.type, is_default=True).update({'is_default': False})
        
        try:
            db.session.add(template)
            db.session.commit()
            return jsonify({'success': True, 'id': template.id})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)})
    
    return render_template('print_template_form.html', template=None)

@app.route('/settings/print-templates/<int:id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
@admin_required
def print_template(id):
    template = PrintTemplate.query.get_or_404(id)
    
    if request.method == 'GET':
        return render_template('print_template_form.html', template=template)
    
    elif request.method == 'PUT':
        data = request.json
        template.name = data['name']
        template.content = data['content']
        template.is_default = data.get('is_default', False)
        
        if template.is_default:
            # Remove default flag from other templates of same type
            PrintTemplate.query.filter(
                PrintTemplate.type == template.type,
                PrintTemplate.id != template.id,
                PrintTemplate.is_default == True
            ).update({'is_default': False})
        
        try:
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)})
    
    elif request.method == 'DELETE':
        if template.is_default:
            return jsonify({'success': False, 'error': 'Cannot delete default template'})
        
        try:
            db.session.delete(template)
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)})

def get_default_template(type):
    template = PrintTemplate.query.filter_by(type=type, is_default=True).first()
    if not template:
        # Create default template if none exists
        if type == 'invoice':
            with open('templates/print_invoice.html', 'r') as f:
                content = f.read()
        else:
            with open('templates/print_summary.html', 'r') as f:
                content = f.read()
        
        template = PrintTemplate(
            name=f'Default {type.title()} Template',
            type=type,
            content=content,
            is_default=True
        )
        db.session.add(template)
        db.session.commit()
    
    return template

@app.route('/settings')
@login_required
@permission_required('manage_settings')
def settings():
    # Get print templates
    invoice_templates = PrintTemplate.query.filter_by(type='invoice').all()
    summary_templates = PrintTemplate.query.filter_by(type='summary').all()
    settings = Settings.query.first()
    current_wallpaper = url_for('static', filename=f'wallpapers/{settings.wallpaper_path}') if settings and settings.wallpaper_path else None
    
    return render_template('settings.html', 
                         invoice_templates=invoice_templates,
                         summary_templates=summary_templates,
                         current_wallpaper=current_wallpaper)

@app.route('/settings/upload-wallpaper', methods=['POST'])
@login_required
@admin_required
def upload_wallpaper():
    if 'wallpaper' not in request.files:
        flash('No file uploaded', 'error')
        return redirect(url_for('settings'))
    
    file = request.files['wallpaper']
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('settings'))
    
    if file and file.filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
        # Create wallpapers directory if it doesn't exist
        wallpapers_dir = os.path.join(app.static_folder, 'wallpapers')
        if not os.path.exists(wallpapers_dir):
            os.makedirs(wallpapers_dir)
        
        # Save the file
        filename = f'wallpaper_{datetime.now().strftime("%Y%m%d_%H%M%S")}{os.path.splitext(file.filename)[1]}'
        file.save(os.path.join(wallpapers_dir, filename))
        
        # Update settings
        settings = Settings.query.first()
        if not settings:
            settings = Settings()
            db.session.add(settings)
        
        # Remove old wallpaper if it exists
        if settings.wallpaper_path:
            old_wallpaper_path = os.path.join(wallpapers_dir, settings.wallpaper_path)
            if os.path.exists(old_wallpaper_path):
                os.remove(old_wallpaper_path)
        
        settings.wallpaper_path = filename
        db.session.commit()
        
        flash('Wallpaper updated successfully', 'success')
    else:
        flash('Invalid file type. Please upload an image file.', 'error')
    
    return redirect(url_for('settings'))

@app.route('/settings/remove-wallpaper', methods=['POST'])
@login_required
@admin_required
def remove_wallpaper():
    settings = Settings.query.first()
    if settings and settings.wallpaper_path:
        # Remove the file
        wallpaper_path = os.path.join(app.static_folder, 'wallpapers', settings.wallpaper_path)
        if os.path.exists(wallpaper_path):
            os.remove(wallpaper_path)
        
        # Clear the path in settings
        settings.wallpaper_path = None
        db.session.commit()
        
        flash('Wallpaper removed successfully', 'success')
    
    return redirect(url_for('settings'))

@app.route('/settings/calculator-code', methods=['GET', 'POST'])
@login_required
@admin_required
def calculator_code():
    if request.method == 'POST':
        try:
            data = request.get_json()
            code = data.get('code')
            if not code:
                return jsonify({'error': 'Code is required'}), 400
                
            settings = Settings.query.first()
            if not settings:
                settings = Settings()
                db.session.add(settings)
            
            settings.calculator_code = code
            print(f"Updating calculator code to: {code}")  # Debug log
            
            try:
                db.session.commit()
                print("Database commit successful")  # Debug log
                return jsonify({'success': True, 'code': settings.calculator_code})
            except Exception as e:
                print(f"Database commit failed: {str(e)}")  # Debug log
                db.session.rollback()
                return jsonify({'error': f'Failed to save code: {str(e)}'}), 500
                
        except Exception as e:
            print(f"Error in calculator code update: {str(e)}")  # Debug log
            return jsonify({'error': str(e)}), 500
    else:
        try:
            settings = Settings.query.first()
            code = settings.calculator_code if settings else '9999'
            print(f"Current calculator code: {code}")  # Debug log
            return jsonify({'code': code})
        except Exception as e:
            print(f"Error retrieving calculator code: {str(e)}")  # Debug log
            return jsonify({'error': str(e)}), 500

@app.route('/settings/backup')
@login_required
@admin_required
def backup_data():
    # Get all data from database
    products = Product.query.all()
    invoices = Invoice.query.all()
    invoice_items = InvoiceItem.query.all()
    templates = PrintTemplate.query.all()
    settings = Settings.query.first()
    
    # Create backup data structure
    backup = {
        'products': [{
            'item_code': p.item_code,
            'description': p.description,
            'uom': p.uom,
            'price': p.price,
            'stock': p.stock,
            'restock_level': p.restock_level,
            'tamil_name': p.tamil_name
        } for p in products],
        'invoices': [{
            'order_number': i.order_number,
            'date': i.date.isoformat(),
            'customer_name': i.customer_name,
            'total_amount': i.total_amount,
            'total_items': i.total_items
        } for i in invoices],
        'invoice_items': [{
            'invoice_order_number': ii.invoice.order_number,
            'product_item_code': ii.product.item_code,
            'quantity': ii.quantity,
            'price': ii.price,
            'amount': ii.amount
        } for ii in invoice_items],
        'templates': [{
            'name': t.name,
            'type': t.type,
            'content': t.content,
            'is_default': t.is_default
        } for t in templates],
        'settings': {
            'calculator_code': settings.calculator_code if settings else '9999',
            'wallpaper_path': settings.wallpaper_path if settings and settings.wallpaper_path else None
        }
    }
    
    return jsonify(backup)

@app.route('/settings/restore', methods=['POST'])
@login_required
@admin_required
def restore_data():
    try:
        data = request.get_json()
        
        # Clear existing data
        PrintTemplate.query.delete()
        InvoiceItem.query.delete()
        Invoice.query.delete()
        Product.query.delete()
        Settings.query.delete()
        
        # Restore products
        for product_data in data.get('products', []):
            product = Product(**product_data)
            db.session.add(product)
        
        # Restore invoices
        for invoice_data in data.get('invoices', []):
            date_str = invoice_data.pop('date')
            invoice_data['date'] = datetime.fromisoformat(date_str).date()
            invoice = Invoice(**invoice_data)
            db.session.add(invoice)
            
            # Restore invoice items
        for item_data in data.get('invoice_items', []):
            invoice = Invoice.query.filter_by(order_number=item_data['invoice_order_number']).first()
            product = Product.query.filter_by(item_code=item_data['product_item_code']).first()
            if invoice and product:
                item = InvoiceItem(
                    invoice=invoice,
                    product=product,
                    quantity=item_data['quantity'],
                    price=item_data['price'],
                    amount=item_data['amount']
                )
                db.session.add(item)
        
        # Restore templates
        for template_data in data.get('templates', []):
            template = PrintTemplate(**template_data)
            db.session.add(template)
        
        # Restore settings
        settings_data = data.get('settings', {})
        settings = Settings(
            calculator_code=settings_data.get('calculator_code', '9999'),
            wallpaper_path=settings_data.get('wallpaper_path')
        )
        db.session.add(settings)
        
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/settings/delete-invoices', methods=['POST'])
@login_required
@admin_required
def delete_invoices():
    try:
        data = request.json
        restore_stock = data.get('restore_stock', False)
        _delete_all_invoices(restore_stock)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/settings/delete-products', methods=['POST'])
@login_required
@admin_required
def delete_products():
    try:
        _delete_all_products()
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/settings/delete-all', methods=['POST'])
@login_required
@admin_required
def delete_all():
    try:
        _delete_all_products()  # This also deletes invoices
        PrintTemplate.query.delete()
        CustomerTransaction.query.delete()
        CustomerReceivable.query.delete()
        Customer.query.delete()
        settings = Settings.query.first()
        if settings:
            settings.calculator_code = '9999'
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/settings/print-templates/<int:id>', methods=['DELETE'])
@login_required
@admin_required
def delete_print_template(id):
    template = PrintTemplate.query.get_or_404(id)
    if template.is_default:
        return jsonify({'success': False, 'error': 'Cannot delete default template'})
    
    try:
        db.session.delete(template)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/sales_trend')
def sales_trend():
    return jsonify({'labels': [], 'values': []})

@app.route('/api/top_products')
def top_products():
    return jsonify({'labels': [], 'values': []})

@app.route('/api/slow_moving_products')
def slow_moving_products():
    return jsonify({'labels': [], 'values': []})

@app.route('/api/stock_sales_ratio')
def stock_sales_ratio():
    return jsonify({'labels': [], 'values': []})

@app.route('/api/inventory_aging')
def inventory_aging():
    return jsonify({'labels': [], 'values': []})

@app.route('/api/sales_forecast')
def sales_forecast():
    return jsonify({'labels': [], 'values': []})

@app.route('/api/sales_performance')
@login_required
def sales_performance():
    return jsonify({'labels': [], 'values': []})

@app.route('/api/sales_growth')
@login_required
def sales_growth():
    return jsonify({'labels': [], 'values': []})

@app.route('/api/sales_trend_by_period')
@login_required
def sales_trend_by_period():
    return jsonify({'labels': [], 'values': []})

@app.route('/users/generate-2fa', methods=['POST'])
@login_required
@admin_required
def generate_2fa():
    data = request.get_json()
    username = data.get('username')
    if not username:
        return jsonify({'success': False, 'error': 'Username is required'})
    
    # Generate new TOTP secret
    totp_secret = random_base32()
    totp = TOTP(totp_secret)
    
    # Generate TOTP URI for QR code
    totp_uri = f'otpauth://totp/Inventory:{username}?secret={totp_secret}&issuer=Inventory'
    
    return jsonify({
        'success': True,
        'totp_secret': totp_secret,
        'totp_uri': totp_uri
    })

@app.route('/users/<int:id>/generate-2fa', methods=['POST'])
@login_required
@admin_required
def generate_user_2fa(id):
    user = User.query.get_or_404(id)
    data = request.get_json()
    username = data.get('username', user.username)
    
    # Generate new TOTP secret
    totp_secret = random_base32()
    totp = TOTP(totp_secret)
    
    # Save the TOTP secret to the user
    try:
        user.totp_secret = totp_secret
        db.session.commit()
        
        # Generate TOTP URI for QR code
        totp_uri = f'otpauth://totp/Inventory:{username}?secret={totp_secret}&issuer=Inventory'
        
        return jsonify({
            'success': True,
            'totp_secret': totp_secret,
            'totp_uri': totp_uri
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/invoices/list')
def list_invoices():
    invoices = Invoice.query.order_by(Invoice.date.desc()).all()
    return jsonify({
        'invoices': [
            {
                'id': invoice.id,
                'order_number': invoice.order_number,
                'date': invoice.date.isoformat(),
                'customer_name': invoice.customer_name,
                'total_amount': invoice.total_amount
            } for invoice in invoices
        ]
    })

@app.route('/reports')
@login_required
@permission_required('view_reports')
def reports():
    return redirect(url_for('index'))

@app.route('/customers')
@login_required
@permission_required('view_customers')
def customers():
    customers = Customer.query.all()
    return render_template('customers.html', customers=customers)

@app.route('/customers', methods=['POST'])
@login_required
def add_customer():
    data = request.json
    
    # Server-side validation
    if not data.get('name'):
        return jsonify({'success': False, 'error': 'Name is required'})
    
    customer = Customer(
        name=data['name'],
        phone=data.get('phone'),
        email=data.get('email'),
        address=data.get('address')
    )
    
    try:
        db.session.add(customer)
        db.session.commit()
        return jsonify({'success': True, 'id': customer.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/customers/<int:id>')
@login_required
def customer_detail(id):
    customer = Customer.query.get_or_404(id)
    transactions = CustomerTransaction.query.filter_by(customer_id=id).order_by(CustomerTransaction.date.desc()).all()
    return render_template('customer_detail.html', customer=customer, transactions=transactions)

@app.route('/customers/<int:id>', methods=['PUT', 'DELETE'])
@login_required
def update_customer(id):
    customer = Customer.query.get_or_404(id)
    
    if request.method == 'PUT':
        data = request.json
        
        # Server-side validation
        if not data.get('name'):
            return jsonify({'success': False, 'error': 'Name is required'})
        
        try:
            customer.name = data['name']
            customer.phone = data.get('phone')
            customer.email = data.get('email')
            customer.address = data.get('address')
            
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)})
    
    elif request.method == 'DELETE':
        try:
            # Check if customer has any invoices
            if customer.invoices:
                return jsonify({'success': False, 'error': 'Cannot delete customer with existing invoices'})
            
            # Delete all transactions and receivables
            CustomerTransaction.query.filter_by(customer_id=id).delete()
            CustomerReceivable.query.filter_by(customer_id=id).delete()
            
            # Delete the customer
            db.session.delete(customer)
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)})

@app.route('/customers/payment', methods=['POST'])
@login_required
def add_customer_payment():
    data = request.json
    
    # Server-side validation
    if not all(key in data for key in ('customer_id', 'amount', 'payment_method')):
        return jsonify({'success': False, 'error': 'Missing required fields'})
    
    try:
        amount = float(data['amount'])
        customer_id = int(data['customer_id'])
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid amount or customer ID'})
    
    customer = Customer.query.get_or_404(customer_id)
    
    transaction = CustomerTransaction(
        customer_id=customer_id,
        amount=amount,
        transaction_type='payment',
        payment_method=data['payment_method'],
        reference_number=data.get('reference_number'),
        notes=data.get('notes')
    )
    
    try:
        db.session.add(transaction)
        customer.update_balance()
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/customers/<int:id>/link-invoice', methods=['POST'])
@login_required
def link_invoice_to_customer(id):
    try:
        data = request.json
        invoice_id = data.get('invoice_id')
        additional_amount = float(data.get('additional_amount', 0))
        notes = data.get('notes', '')
        
        # Get the invoice and customer
        invoice = Invoice.query.get_or_404(invoice_id)
        customer = Customer.query.get_or_404(id)
        
        # Link the invoice to the customer
        invoice.customer_id = id
        invoice.customer_name = customer.name
        
        # Create a receivable record
        receivable = CustomerReceivable(
            customer_id=id,
            amount=invoice.total_amount,
            notes=notes,
            invoice_id=invoice_id,
            additional_amount=additional_amount
        )
        
        db.session.add(receivable)
        
        # Update customer balance
        customer.update_balance()
        
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/customers/<int:id>/receivable', methods=['POST'])
@login_required
def add_customer_receivable(id):
    try:
        data = request.json
        amount = float(data.get('amount'))
        notes = data.get('notes')
        
        if not amount or not notes:
            return jsonify({'success': False, 'error': 'Amount and notes are required'})
        
        customer = Customer.query.get_or_404(id)
        
        receivable = CustomerReceivable(
            customer_id=id,
            amount=amount,
            notes=notes
        )
        
        db.session.add(receivable)
        customer.update_balance()
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/customers/transaction/<int:id>', methods=['PUT', 'DELETE'])
@login_required
def manage_transaction(id):
    transaction = CustomerTransaction.query.get_or_404(id)
    
    if request.method == 'PUT':
        data = request.json
        try:
            transaction.amount = float(data['amount'])
            transaction.payment_method = data['payment_method']
            transaction.reference_number = data.get('reference_number')
            transaction.notes = data.get('notes')
            
            # Update customer balance
            transaction.customer.update_balance()
            
            # Update invoice payment status if this transaction affects any invoices
            for invoice in transaction.customer.invoices:
                _update_invoice_status(invoice)
            
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)})
    
    elif request.method == 'DELETE':
        try:
            # Update customer balance before deleting
            customer = transaction.customer
            db.session.delete(transaction)
            
            # Update customer balance
            customer.update_balance()
            
            # Update invoice payment status for all customer invoices
            for invoice in customer.invoices:
                _update_invoice_status(invoice)
            
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)})

def _update_invoice_status(invoice):
    """Helper function to update invoice payment status based on payments and receivables"""
    if not invoice.customer:
        return
        
    # Get all customer payments
    total_payments = sum(
        transaction.amount 
        for transaction in invoice.customer.transactions 
        if transaction.transaction_type == 'payment'
    )
    
    # Get all receivables
    total_receivables = sum(
        receivable.amount + receivable.additional_amount 
        for receivable in invoice.customer.receivables
    )
    
    # Calculate total amount due
    total_due = total_receivables
    
    if total_payments >= total_due:
        invoice.payment_status = 'paid'
    elif total_payments > 0:
        invoice.payment_status = 'partial'
    else:
        invoice.payment_status = 'pending'

@app.route('/customers/receivable/<int:id>', methods=['PUT', 'DELETE'])
@login_required
def manage_receivable(id):
    receivable = CustomerReceivable.query.get_or_404(id)
    
    if request.method == 'PUT':
        data = request.json
        try:
            receivable.amount = float(data['amount'])
            receivable.additional_amount = float(data.get('additional_amount', 0))
            receivable.notes = data['notes']
            
            # Update customer balance
            receivable.customer.update_balance()
            
            # Update invoice status
            if receivable.invoice:
                _update_invoice_status(receivable.invoice)
            
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)})
    
    elif request.method == 'DELETE':
        try:
            # Store reference to customer and invoice before deleting
            customer = receivable.customer
            invoice = receivable.invoice
            
            # If this receivable is linked to an invoice, unlink it
            if invoice:
                invoice.customer_id = None
                invoice.customer_name = None
                invoice.payment_status = 'pending'
            
            # Delete the receivable
            db.session.delete(receivable)
            
            # Update customer balance
            customer.update_balance()
            
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)})

# Models
class Supplier(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    phone = db.Column(db.String(20))
    email = db.Column(db.String(120))
    address = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    products = db.relationship('Product', secondary='supplier_products', backref='suppliers')

# Association table for supplier-product relationship
supplier_products = db.Table('supplier_products',
    db.Column('supplier_id', db.Integer, db.ForeignKey('supplier.id'), primary_key=True),
    db.Column('product_id', db.Integer, db.ForeignKey('product.id'), primary_key=True)
)

# Add supplier routes
@app.route('/suppliers')
@login_required
def suppliers():
    suppliers = Supplier.query.all()
    return render_template('suppliers.html', suppliers=suppliers)

@app.route('/suppliers', methods=['POST'])
@login_required
def add_supplier():
    data = request.json
    
    # Server-side validation
    if not data.get('name'):
        return jsonify({'success': False, 'error': 'Name is required'})
    
    supplier = Supplier(
        name=data['name'],
        phone=data.get('phone'),
        email=data.get('email'),
        address=data.get('address')
    )
    
    try:
        db.session.add(supplier)
        db.session.commit()
        return jsonify({'success': True, 'id': supplier.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/suppliers/<int:id>')
@login_required
def supplier_detail(id):
    supplier = Supplier.query.get_or_404(id)
    products = Product.query.all()  # Get all products for linking
    
    # Get sort parameter from request
    sort_by = request.args.get('sort', '')
    
    # Define stock level thresholds
    LOW_STOCK = 10
    MEDIUM_STOCK = 30
    
    if sort_by == 'low_stock':
        supplier.products.sort(key=lambda x: x.stock)
        products.sort(key=lambda x: x.stock)
    elif sort_by == 'medium_stock':
        supplier.products.sort(key=lambda x: abs(x.stock - MEDIUM_STOCK))
        products.sort(key=lambda x: abs(x.stock - MEDIUM_STOCK))
    elif sort_by == 'high_stock':
        supplier.products.sort(key=lambda x: -x.stock)
        products.sort(key=lambda x: -x.stock)
    
    return render_template('supplier_detail.html', 
                         supplier=supplier, 
                         products=products,
                         low_stock=LOW_STOCK,
                         medium_stock=MEDIUM_STOCK)

@app.route('/suppliers/<int:id>', methods=['PUT'])
@login_required
def update_supplier(id):
    supplier = Supplier.query.get_or_404(id)
    data = request.json
    
    # Server-side validation
    if not data.get('name'):
        return jsonify({'success': False, 'error': 'Name is required'})
    
    try:
        supplier.name = data['name']
        supplier.phone = data.get('phone')
        supplier.email = data.get('email')
        supplier.address = data.get('address')
        
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/suppliers/<int:id>/products', methods=['POST'])
@login_required
def link_supplier_products(id):
    supplier = Supplier.query.get_or_404(id)
    data = request.json
    
    try:
        # Clear existing products
        supplier.products = []
        
        # Add selected products
        for product_id in data.get('product_ids', []):
            product = Product.query.get(product_id)
            if product:
                supplier.products.append(product)
        
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/suppliers/<int:id>', methods=['DELETE'])
@login_required
def delete_supplier(id):
    supplier = Supplier.query.get_or_404(id)
    try:
        db.session.delete(supplier)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/suppliers/<int:id>/products/import', methods=['POST'])
@login_required
def import_supplier_products(id):
    supplier = Supplier.query.get_or_404(id)
    data = request.json
    
    if not data or 'product_codes' not in data:
        return jsonify({'success': False, 'error': 'No product codes provided'})
    
    try:
        # Get all products with matching codes
        products = Product.query.filter(Product.item_code.in_(data['product_codes'])).all()
        
        # Add products to supplier
        for product in products:
            if product not in supplier.products:
                supplier.products.append(product)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Successfully linked {len(products)} products to supplier'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/suppliers/<int:id>/products/export')
@login_required
def export_supplier_products(id):
    supplier = Supplier.query.get_or_404(id)
    
    # Create DataFrame with supplier's products
    df = pd.DataFrame([{
        'Item Code': p.item_code,
        'Description': p.description,
        'Tamil Name': p.tamil_name,
        'UOM': p.uom,
        'Price': p.price,
        'Stock': p.stock,
        'Restock Level': p.restock_level,
        'Stock Locations': p.stock_locations,
        'Tags': p.tags,
        'Notes': p.notes
    } for p in supplier.products])

    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    workbook = writer.book
    
    # Add formats
    header_format = workbook.add_format({
        'bold': True,
        'text_wrap': True,
        'valign': 'top',
        'bg_color': '#D9D9D9',
        'border': 1
    })
    
    # Write DataFrame to Excel
    df.to_excel(writer, sheet_name='Products', index=False)
    
    # Get the worksheet object
    worksheet = writer.sheets['Products']
    
    # Format the header row
    for col_num, value in enumerate(df.columns.values):
        worksheet.write(0, col_num, value, header_format)
        worksheet.set_column(col_num, col_num, 15)  # Set column width
    
    # Set specific column widths
    worksheet.set_column('B:B', 30)  # Description column
    worksheet.set_column('C:C', 20)  # Tamil Name column
    worksheet.set_column('H:H', 25)  # Stock Locations column
    worksheet.set_column('I:I', 20)  # Tags column
    worksheet.set_column('J:J', 30)  # Notes column
    
    writer.close()
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{supplier.name}_products.xlsx'
    )

@app.route('/save_invoice', methods=['POST'])
def save_invoice():
    # ... existing invoice saving code ...
    
    # Change the redirect from /invoices to /new_invoice
    return redirect(url_for('new_invoice'))

# Add this function for database connection
def get_db_connection():
    import psycopg2
    from psycopg2.extras import DictCursor
    return psycopg2.connect(
        os.getenv('DATABASE_URL'),
        cursor_factory=DictCursor
    )

@app.route('/offline.html')
def offline():
    return render_template('offline.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        try:
            username = request.form.get('username')
            password = request.form.get('password')
            totp_token = request.form.get('totp_token')
            
            print(f"Login attempt for user: {username}")
            
            user = User.query.filter_by(username=username).first()
            if not user:
                print(f"User not found: {username}")
                flash('Invalid username or password', 'error')
                return redirect(url_for('login'))
            
            # Verify password using the same method used for hashing
            if not check_password_hash(user.password, password):
                print(f"Invalid password for user: {username}")
                flash('Invalid username or password', 'error')
                return redirect(url_for('login'))
            
            # Check 2FA if enabled
            if user.totp_enabled:
                if not totp_token:
                    flash('2FA token required', 'error')
                    return redirect(url_for('login'))
                
                if not user.verify_totp(totp_token):
                    print(f"Invalid 2FA token for user: {username}")
                    flash('Invalid 2FA token', 'error')
                    return redirect(url_for('login'))
            
            # Login successful
            session.clear()
            session['user_id'] = user.id
            session['logged_in'] = True
            session['is_admin'] = user.role == 'admin'
            
            # Store user permissions in session
            if user.role == 'admin':
                session['permissions'] = [p.name for p in Permission.query.all()]
            else:
                session['permissions'] = [p.name for p in user.permissions]
            
            print(f"Login successful for user: {username}")
            print(f"User permissions: {session['permissions']}")  # Debug log
            return redirect(url_for('index'))
            
        except Exception as e:
            print(f"Login error for user {username if username else 'unknown'}: {str(e)}")
            print(f"Full traceback: {traceback.format_exc()}")
            flash('An error occurred during login. Please try again.', 'error')
            return redirect(url_for('login'))
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.before_request
def require_login():
    # List of routes that don't require login
    public_routes = ['login', 'static']
    
    # Check if the requested endpoint is in public routes
    if request.endpoint and request.endpoint not in public_routes:
        if 'logged_in' not in session:
            return redirect(url_for('login'))

# Users management routes
@app.route('/users')
@login_required
@admin_required
def users():
    try:
        current_user = User.query.get(session.get('user_id'))
        if not current_user:
            flash('Session expired. Please login again.', 'error')
            return redirect(url_for('login'))
        
        if current_user.role != 'admin':
            flash('Access denied. Admin privileges required.', 'error')
            return redirect(url_for('index'))
        
        users = User.query.all()
        permissions = Permission.query.all()
        
        # Group permissions by category for better organization
        grouped_permissions = {
            'products': [],
            'invoices': [],
            'customers': [],
            'suppliers': [],
            'others': []
        }
        
        for perm in permissions:
            if perm.name.startswith(('view_product', 'edit_product')):
                grouped_permissions['products'].append(perm)
            elif 'invoice' in perm.name:
                grouped_permissions['invoices'].append(perm)
            elif 'customer' in perm.name:
                grouped_permissions['customers'].append(perm)
            elif 'supplier' in perm.name:
                grouped_permissions['suppliers'].append(perm)
            else:
                grouped_permissions['others'].append(perm)
        
        return render_template(
            'users.html',
            users=users,
            permissions=permissions,
            grouped_permissions=grouped_permissions,
            current_user=current_user
        )
    except Exception as e:
        print(f"Error in users route: {str(e)}")
        print(f"Full traceback: {traceback.format_exc()}")  # Add full traceback
        db.session.rollback()
        flash('An error occurred while loading users. Please try again.', 'error')
        return redirect(url_for('index'))

@app.route('/users', methods=['POST'])
@login_required
@admin_required
def add_user():
    try:
        data = request.get_json()
        
        # Validate required fields
        if not all(key in data for key in ['username', 'password', 'role']):
            return jsonify({'error': 'Missing required fields'}), 400
            
        # Check if username exists
        if User.query.filter_by(username=data['username']).first():
            return jsonify({'error': 'Username already exists'}), 400
        
        # Create new user
        new_user = User(
            username=data['username'],
            password=generate_password_hash(data['password'], method='pbkdf2:sha256'),
            role=data['role'],
            totp_enabled=data.get('totp_enabled', False)
        )
        
        # Handle permissions
        if data['role'] != 'admin' and 'permissions' in data:
            try:
                permissions = Permission.query.filter(Permission.id.in_(data['permissions'])).all()
                new_user.permissions = permissions
            except Exception as e:
                print(f"Error setting permissions: {str(e)}")
                return jsonify({'error': 'Invalid permissions data'}), 400
        
        db.session.add(new_user)
        db.session.commit()
        
        print(f"User created successfully: {new_user.username}")
        return jsonify({'success': True})
        
    except Exception as e:
        print(f"Error creating user: {str(e)}")
        print(f"Full traceback: {traceback.format_exc()}")
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

def update_session_permissions(user_id):
    """Update the session permissions for a user"""
    user = User.query.get(user_id)
    if user:
        if user.role == 'admin':
            session['permissions'] = [p.name for p in Permission.query.all()]
        else:
            session['permissions'] = [p.name for p in user.permissions]
        session['is_admin'] = user.role == 'admin'

@app.route('/users/<int:id>', methods=['PUT'])
@login_required
@admin_required
def update_user(id):
    user = User.query.get_or_404(id)
    data = request.get_json()
    
    if data['username'] != user.username and User.query.filter_by(username=data['username']).first():
        return jsonify({'error': 'Username already exists'}), 400
    
    try:
        user.username = data['username']
        user.role = data['role']
        if data.get('password'):
            user.password = generate_password_hash(data['password'])
        
        # Update 2FA status
        user.totp_enabled = data.get('totp_enabled', False)
        if not user.totp_enabled:
            user.totp_secret = None  # Clear TOTP secret if 2FA is disabled
        
        # Update permissions if not admin
        if data['role'] != 'admin':
            permissions = Permission.query.filter(Permission.id.in_(data.get('permissions', []))).all()
            user.permissions = permissions
        else:
            user.permissions = []  # Clear permissions for admin as they have all permissions
        
        db.session.commit()
        
        # Update session permissions if the updated user is the current user
        if id == session.get('user_id'):
            update_session_permissions(id)
            
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/users/<int:id>', methods=['DELETE'])
@login_required
@admin_required
def delete_user(id):
    if id == session.get('user_id'):
        return jsonify({'error': 'Cannot delete your own account'}), 400
    
    user = User.query.get_or_404(id)
    try:
        db.session.delete(user)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/users/<int:id>/permissions')
@login_required
@admin_required
def get_user_permissions(id):
    user = User.query.get_or_404(id)
    return jsonify({
        'permissions': [p.id for p in user.permissions]
    })

# Add error handlers
@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return "Internal Server Error. Please try again.", 500

@app.errorhandler(504)
def gateway_timeout(error):
    return "Request timed out. Please try again.", 504

# Add database connection timeout and pool settings
app.config['SQLALCHEMY_POOL_SIZE'] = 10
app.config['SQLALCHEMY_POOL_TIMEOUT'] = 30
app.config['SQLALCHEMY_POOL_RECYCLE'] = 1800  # Recycle connections after 30 minutes

@app.route('/sw.js')
def service_worker():
    return '', 404

@app.route('/manifest.json')
def manifest():
    return '', 404

if __name__ == '__main__':
    try:
        init_db()  # Initialize database with sample data
        app.run(host='0.0.0.0', port=5000, debug=True)
    except Exception as e:
        print(f"Error starting application: {str(e)}")
        raise